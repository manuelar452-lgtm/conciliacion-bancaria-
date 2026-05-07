from __future__ import annotations
import os
import re
import shutil
import tempfile
import numpy as np
import pandas as pd
from datetime import date, datetime, timedelta
import fitz          # PyMuPDF: renderiza PDF→imagen sin poppler
import pytesseract   # OCR sobre esas imágenes
from PIL import Image
import io

# Buscar tesseract en paths comunes del sistema
for _p in ("/usr/bin/tesseract", "/usr/local/bin/tesseract"):
    if os.path.isfile(_p):
        pytesseract.pytesseract.tesseract_cmd = _p
        break

import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(
    page_title="Conciliación Bancaria",
    page_icon="🏦",
    layout="centered",
)

col_a, col_b = st.columns(2)
uploaded_ext = col_a.file_uploader("📄 Extracto bancario (PDF)", type="pdf")
uploaded_aux = col_b.file_uploader("📒 Auxiliar contabilidad (PDF)", type="pdf")

if not uploaded_ext or not uploaded_aux:
    st.info("Sube los dos archivos PDF para continuar.")
    st.stop()

col_a.success(uploaded_ext.name)
col_b.success(uploaded_aux.name)

if st.button("Generar conciliación", type="primary"):

    with st.spinner("Procesando... (puede tardar 1-2 minutos)"):

        # ── Guardar PDFs temporales ───────────────────────────────────────────
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
            tmp.write(uploaded_ext.read())
            pdf_ext_path = tmp.name
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
            tmp.write(uploaded_aux.read())
            pdf_aux_path = tmp.name

        # ── Utilidades ────────────────────────────────────────────────────────

        def _parse_num(text):
            clean = re.sub(r"[^\d.]", "", str(text))
            try:    return float(clean) if clean else None
            except: return None

        def _is_currency(text):
            t = re.sub(r"[\[\|\]]+", "", text).strip()
            return bool(
                re.match(r"^-?\d{1,3}(,\d{3})+(\.\d{2})?$", t) or
                re.match(r"^-?\d{4,}(\.\d{2})?$", t)
            )

        def _fmt_cop(v):
            if v is None or (isinstance(v, float) and np.isnan(v)): return ""
            try:    return f"$ {float(v):>20,.0f}"
            except: return str(v)

        def _plumber_lines(page):
            """Extrae líneas desde pdfplumber (sin OCR). Retorna (lines, page_width)."""
            words = page.extract_words(x_tolerance=3, y_tolerance=3,
                                       keep_blank_chars=False)
            if not words:
                return [], page.width
            lines, cur, cur_y = [], [], None
            for w in sorted(words, key=lambda w: (round(w["top"], 1), w["x0"])):
                if cur_y is None or abs(w["top"] - cur_y) <= 4:
                    cur.append((w["x0"], w["text"]))
                    if cur_y is None: cur_y = w["top"]
                else:
                    if cur: lines.append(sorted(cur))
                    cur, cur_y = [(w["x0"], w["text"])], w["top"]
            if cur: lines.append(sorted(cur))
            return lines, page.width

        def _fitz_lines(pdf_path, page_num=0):
            """Extrae líneas de texto de una página usando PyMuPDF (sin binarios del sistema)."""
            doc = fitz.open(pdf_path)
            page = doc[page_num]
            words = page.get_text("words")  # (x0,y0,x1,y1,word,block,line,word_idx)
            doc.close()
            if not words:
                return [], page.rect.width
            w = page.rect.width
            # Agrupar por Y (tolerancia 4pt)
            lines, cur, cur_y = [], [], None
            for x0, y0, x1, y1, word, *_ in sorted(words, key=lambda r: (round(r[1], 1), r[0])):
                if cur_y is None or abs(y0 - cur_y) <= 4:
                    cur.append((x0, word))
                    if cur_y is None: cur_y = y0
                else:
                    if cur: lines.append(sorted(cur))
                    cur, cur_y = [(x0, word)], y0
            if cur: lines.append(sorted(cur))
            return lines, w

        def _fitz_all_lines(pdf_path):
            """Extrae líneas de todas las páginas con PyMuPDF."""
            doc = fitz.open(pdf_path)
            all_lines = []
            widths = []
            for page in doc:
                words = page.get_text("words")
                w = page.rect.width
                widths.append(w)
                lines, cur, cur_y = [], [], None
                for x0, y0, x1, y1, word, *_ in sorted(words, key=lambda r: (round(r[1],1), r[0])):
                    if cur_y is None or abs(y0 - cur_y) <= 4:
                        cur.append((x0, word))
                        if cur_y is None: cur_y = y0
                    else:
                        if cur: lines.append(sorted(cur))
                        cur, cur_y = [(x0, word)], y0
                if cur: lines.append(sorted(cur))
                all_lines.append((lines, w))
            doc.close()
            return all_lines

        # ── Extracto: parser genérico (Itau, Banco de Bogotá, etc.) ─────────────
        def _clean(t: str) -> str:
            """Elimina artefactos OCR como | al inicio/fin."""
            return re.sub(r"[\[\|\]]+", "", t).strip()

        def _is_amount(text: str) -> bool:
            t = _clean(text)
            return bool(
                re.match(r"^-?\d{1,3}(,\d{3})+(\.\d{1,2})?$", t) or   # US: 1,334,847.00
                re.match(r"^-?\d{4,}\.\d{2}$",                  t) or   # sin comas: 1334847.00
                re.match(r"^-?\d{1,3}(\.\d{3})+(,\d{1,2})?$",  t) or   # EUR: 1.334.847,00
                re.match(r"^-?\d{6,}$",                          t)      # entero sin formato
            )

        def _to_float(text: str):
            t = _clean(text)
            # Formato europeo: puntos como miles, coma como decimal → 1.334.847,00
            if re.match(r"^-?\d{1,3}(\.\d{3})+(,\d{1,2})?$", t):
                t = t.replace(".", "").replace(",", ".")
            else:
                t = t.replace(",", "")
            try:    return float(t)
            except: return None

        def _parse_extracto_page(lines, w):
            rows = []
            for line in lines:
                if not line: continue

                # ── Detectar fecha (primeros tokens, zona flexible hasta 25%) ───
                day, mon = None, None
                for tok_x, tok_t in line:
                    if tok_x >= w * 0.25:
                        break
                    tc = _clean(tok_t)
                    m = re.match(r"^(\d{1,2})[/\-](\d{2})$", tc)
                    if m:
                        d, mo = int(m.group(1)), int(m.group(2))
                        if 1 <= d <= 31 and 1 <= mo <= 12:
                            day, mon = d, mo; break
                    if re.match(r"^\d{1,2}$", tc):
                        d = int(tc)
                        if 1 <= d <= 31:
                            day = d; break

                if day is None: continue

                # ── Recopilar todos los montos en la mitad derecha ─────────────
                # Estrategia: tomar todos los números monetarios (x > 45% del ancho),
                # ordenarlos de izquierda a derecha.
                # El más a la derecha = saldo; el anterior = monto de transacción.
                all_amounts = []   # (x, valor)
                i = 0
                while i < len(line):
                    x, t = line[i]
                    tc = _clean(t)
                    if x < w * 0.45:
                        i += 1; continue
                    # Monto negativo partido: "-" seguido de número
                    if tc == "-" and i + 1 < len(line):
                        nx, nt = line[i+1]
                        if _is_amount(nt) and nx >= w * 0.45 and (nx - x) < w * 0.08:
                            v = _to_float(nt)
                            if v: all_amounts.append((nx, -abs(v)))
                            i += 2; continue
                    if _is_amount(tc):
                        v = _to_float(tc)
                        if v is not None and abs(v) >= 50:
                            all_amounts.append((x, v))
                    i += 1

                if len(all_amounts) < 2: continue

                all_amounts.sort(key=lambda a: a[0])
                saldo = abs(all_amounts[-1][1])
                valor = all_amounts[-2][1]

                # ── Clasificar débito / crédito ────────────────────────────────
                desc_tokens = [t for x, t in line if not _is_amount(t) and x < all_amounts[-2][0]]
                desc_lower  = " ".join(desc_tokens).lower()
                cargo_kw = {"cargo", "giro", "gravamen", " db ", "pago",
                            "impuesto", "cheque", "comision", "débito", "debito"}
                if valor < 0:
                    debito, credito = abs(valor), 0.0
                elif any(k in desc_lower for k in cargo_kw):
                    debito, credito = valor, 0.0
                else:
                    debito, credito = 0.0, valor

                # ── Descripción: tokens no-numéricos entre fecha y primer monto ─
                first_amount_x = all_amounts[0][0]
                desc_parts = [
                    _clean(t) for x, t in line
                    if w * 0.08 < x < first_amount_x - w * 0.01
                    and not _is_amount(t)
                    and len(_clean(t)) > 1
                ]
                desc = " ".join(desc_parts[:12])

                rows.append({
                    "dia": day, "mes": mon, "descripcion": desc,
                    "debito": debito, "credito": credito, "saldo": saldo,
                })
            return rows

        def _saldo_final_extracto(images) -> float | None:
            """Busca 'Saldo Final' en el texto del extracto, desde la última página."""
            for img in reversed(images):
                text = pytesseract.image_to_string(img, lang="spa", config="--psm 6")
                for line in text.split("\n"):
                    if re.search(r"saldo\s+(final|al\s+corte|bancario)", line, re.IGNORECASE):
                        nums = re.findall(r"[\d,]+\.\d{2}", line)
                        for n in nums:
                            v = _to_float(n)
                            if v and v > 0:
                                return v
            return None

        def _find_saldo_fitz(pdf_path: str) -> float | None:
            """OCR todas las páginas del PDF para encontrar el Saldo Final."""
            try:
                doc = fitz.open(pdf_path)
                n_pages = len(doc)
                for pg_idx in range(n_pages):
                    page = doc[pg_idx]
                    img  = _fitz_render_page(page, dpi=200)
                    text = pytesseract.image_to_string(img, config="--psm 6")
                    for line in text.split("\n"):
                        if re.search(
                            r"saldo\s*(final|al\s*corte|bancario|disponible)",
                            line, re.IGNORECASE
                        ):
                            # Extraer todos los números de la línea (con o sin decimal)
                            nums = re.findall(r"[\d]{1,3}(?:[,.\s]\d{3})+(?:[,.]\d{2})?|\d{6,}", line)
                            for n_str in reversed(nums):
                                # Normalizar: quitar separadores de miles, manejar decimal
                                cleaned = re.sub(r"[,.\s](?=\d{3}(?:[,.\s]|$))", "", n_str)
                                # Intentar con coma como decimal también
                                cleaned = re.sub(r",(\d{2})$", r".\1", cleaned)
                                try:
                                    v = float(cleaned.replace(",", "").replace(" ", ""))
                                    if v > 100_000:
                                        doc.close()
                                        return v
                                except Exception:
                                    pass
                doc.close()
            except Exception:
                pass
            return None

        def _fitz_render_page(fitz_page, dpi=150):
            """Renderiza una página fitz como imagen PIL (sin poppler)."""
            zoom = dpi / 72
            mat  = fitz.Matrix(zoom, zoom)
            pix  = fitz_page.get_pixmap(matrix=mat, colorspace=fitz.csRGB)
            return Image.frombytes("RGB", [pix.width, pix.height], pix.samples)

        def _ocr_page(fitz_page, dpi=150):
            """OCR sobre una página: renderiza con fitz, lee con tesseract."""
            img   = _fitz_render_page(fitz_page, dpi=dpi)
            data  = pytesseract.image_to_data(
                img, config="--psm 6",
                output_type=pytesseract.Output.DATAFRAME
            )
            data  = data[data.conf > 0].dropna(subset=["text"]).copy()
            data["text"] = data["text"].astype(str).str.strip()
            data  = data[data["text"] != ""].sort_values(["top", "left"])
            lines, cur, cur_y = [], [], None
            for _, row in data.iterrows():
                if cur_y is None or abs(row["top"] - cur_y) <= 12:
                    cur.append((row["left"], row["text"]))
                    cur_y = cur_y or row["top"]
                else:
                    if cur: lines.append(sorted(cur))
                    cur, cur_y = [(row["left"], row["text"])], row["top"]
            if cur: lines.append(sorted(cur))
            return lines, img.width

        def parse_extracto(ano, mes):
            all_rows, filas = [], []
            debug = []
            try:
                doc = fitz.open(pdf_ext_path)
                n   = len(doc)
                for pg, page in enumerate(doc, 1):
                    # Intentar texto directo primero (PDFs digitales)
                    words = page.get_text("words")
                    if words:
                        w = page.rect.width
                        lines, cur, cur_y = [], [], None
                        for x0,y0,x1,y1,word,*_ in sorted(words, key=lambda r:(round(r[1],1),r[0])):
                            if cur_y is None or abs(y0-cur_y)<=4:
                                cur.append((x0,word)); cur_y = cur_y or y0
                            else:
                                if cur: lines.append(sorted(cur))
                                cur, cur_y = [(x0,word)], y0
                        if cur: lines.append(sorted(cur))
                        modo = "fitz"
                    else:
                        # PDF escaneado → OCR
                        lines, w = _ocr_page(page, dpi=200)
                        modo = "ocr"
                    if pg == 1:
                        raw = page.get_text("text") or ""
                        debug = [
                            f"Págs:{n} | Palabras p1:{len(words)} | Motor:{modo}",
                            f"Raw p1: {raw[:200]!r}",
                        ] + [" | ".join(t for _,t in ln) for ln in lines[:6]]
                    rows = _parse_extracto_page(lines, w)
                    filas.append((pg, len(rows), modo))
                    all_rows.extend(rows)
                doc.close()
            except Exception as e:
                debug = [f"ERROR: {e}"]

            st.session_state["plumber_sample"] = debug
            st.session_state["ext_debug"]      = filas

            if not all_rows:
                return pd.DataFrame(), None
            df = pd.DataFrame(all_rows)
            def _make_date(row):
                m = int(row["mes"]) if pd.notna(row.get("mes")) and row.get("mes") else mes
                max_day = 28 if m==2 else 30 if m in [4,6,9,11] else 31
                return date(ano, m, min(int(row["dia"]), max_day))
            df["fecha"] = df.apply(_make_date, axis=1)
            return df.sort_values("fecha").reset_index(drop=True), None

        # ── Informe Movimiento General (última página) ────────────────────────
        # Sin posiciones fijas: usa delta de saldo para determinar débito/crédito
        def parse_informe(img):
            lines = _ocr_lines(img)
            rows = []
            prev_saldo = None
            for line in lines:
                # Línea de asiento: contiene patrón L-1-919-1 etc.
                comprobante = next((t for _, t in line if re.match(r"^[A-Z]-\d+-\d+", t)), None)
                if not comprobante: continue
                fecha_str = next((t for _, t in line if re.match(r"^\d{4}/\d{2}/\d{2}$", t)), None)
                if not fecha_str: continue
                try:    fecha = datetime.strptime(fecha_str, "%Y/%m/%d").date()
                except: continue

                # Montos solo en zona derecha (x > 50%): evita capturar números de descripción
                amounts = sorted(
                    [(x, _parse_num(t)) for x, t in line
                     if _is_currency(t) and _parse_num(t) and _parse_num(t) > 0
                     and x > img.width * 0.50],
                    key=lambda a: a[0]
                )
                if not amounts: continue

                # El saldo es siempre el monto más a la derecha
                saldo = amounts[-1][1]

                # El monto de transacción es el anterior al saldo
                non_saldo = [v for _, v in amounts[:-1]]
                if not non_saldo: continue
                amount = non_saldo[-1]

                # Débito = saldo sube (entra dinero), Crédito = saldo baja (sale dinero)
                if prev_saldo is not None:
                    if saldo >= prev_saldo:
                        debito, credito = amount, 0.0
                    else:
                        debito, credito = 0.0, amount
                else:
                    # Primera fila: usar posición relativa (déb está antes de 68% del ancho)
                    x_amount = amounts[-2][0] if len(amounts) >= 2 else 0
                    if x_amount < img.width * 0.68:
                        debito, credito = amount, 0.0
                    else:
                        debito, credito = 0.0, amount
                prev_saldo = saldo

                # Descripción: tokens entre la fecha y la zona de montos
                fecha_x  = next((x for x, t in line if t == fecha_str), 0)
                monto_x  = amounts[-2][0] if len(amounts) >= 2 else img.width
                desc = " ".join(
                    t for x, t in sorted(line)
                    if x > fecha_x and x < monto_x
                    and not re.match(r"^[\d,\.]+$", t)
                    and t not in ("O", "0", "SUC", "NIT")
                ).strip()

                rows.append({
                    "comprobante": comprobante, "fecha": fecha, "descripcion": desc,
                    "debito": debito, "credito": credito, "saldo": saldo,
                })
            return pd.DataFrame(rows) if rows else pd.DataFrame(
                columns=["comprobante","fecha","descripcion","debito","credito","saldo"])

        def _parse_informe_plumber(page):
            """Auxiliar con pdfplumber: misma lógica de parse_informe pero sin OCR."""
            lines, _ = _plumber_lines(page)
            rows = []
            prev_saldo = None
            for line in lines:
                comprobante = next((t for _, t in line if re.match(r"^[A-Z]-\d+-\d+", t)), None)
                if not comprobante: continue
                fecha_str = next((t for _, t in line if re.match(r"^\d{4}/\d{2}/\d{2}$", t)), None)
                if not fecha_str: continue
                try:    fecha = datetime.strptime(fecha_str, "%Y/%m/%d").date()
                except: continue
                amounts = sorted(
                    [(x, _parse_num(t)) for x, t in line
                     if _is_currency(t) and _parse_num(t) and _parse_num(t) > 0
                     and x > page.width * 0.50],
                    key=lambda a: a[0]
                )
                if not amounts: continue
                saldo = amounts[-1][1]
                non_saldo = [v for _, v in amounts[:-1]]
                if not non_saldo: continue
                amount = non_saldo[-1]
                if prev_saldo is not None:
                    if saldo >= prev_saldo: debito, credito = amount, 0.0
                    else:                   debito, credito = 0.0, amount
                else:
                    x_amt = amounts[-2][0] if len(amounts) >= 2 else 0
                    if x_amt < page.width * 0.68: debito, credito = amount, 0.0
                    else:                         debito, credito = 0.0, amount
                prev_saldo = saldo
                fecha_x = next((x for x, t in line if t == fecha_str), 0)
                monto_x  = amounts[-2][0] if len(amounts) >= 2 else page.width
                desc = " ".join(
                    t for x, t in sorted(line)
                    if x > fecha_x and x < monto_x
                    and not re.match(r"^[\d,\.]+$", t)
                    and t not in ("O", "0", "SUC", "NIT")
                ).strip()
                rows.append({"comprobante": comprobante, "fecha": fecha, "descripcion": desc,
                             "debito": debito, "credito": credito, "saldo": saldo})
            return pd.DataFrame(rows) if rows else pd.DataFrame(
                columns=["comprobante","fecha","descripcion","debito","credito","saldo"])

        def _parse_informe_fitz(lines, w):
            """
            Parsea el auxiliar contable que tiene 3 columnas numéricas:
            DÉBITOS | CRÉDITOS | SALDO  (de izquierda a derecha en la zona derecha).
            El saldo es siempre el más a la derecha; el anterior es CRÉDITO; el
            anterior a ese es DÉBITO (si existe).
            """
            rows = []
            for line in lines:
                comprobante = next((t for _, t in line if re.match(r"^[A-Z]-\d+-\d+", t)), None)
                if not comprobante: continue
                fecha_str = next((t for _, t in line if re.match(r"^\d{4}/\d{2}/\d{2}$", t)), None)
                if not fecha_str: continue
                try:    fecha = datetime.strptime(fecha_str, "%Y/%m/%d").date()
                except: continue

                # Recoger todos los montos en la mitad derecha, ordenados por X
                amounts = sorted(
                    [(x, _to_float(t)) for x, t in line
                     if _is_amount(t) and _to_float(t) and abs(_to_float(t)) > 0
                     and x > w * 0.45],
                    key=lambda a: a[0]
                )
                if not amounts: continue

                # Rightmost = SALDO; los anteriores = DÉBITOS | CRÉDITOS según posición X.
                # Con 3+ montos el orden posicional ya distingue las columnas.
                # Con exactamente 2 montos usamos el gap X entre el monto y el SALDO:
                #   gap grande (> 20 % del ancho) → monto está en columna DÉBITO (más a la izq.)
                #   gap pequeño                   → monto está en columna CRÉDITO (adyacente al SALDO)
                saldo = abs(amounts[-1][1])
                x_saldo = amounts[-1][0]

                if len(amounts) >= 3:
                    credito = abs(amounts[-2][1])
                    debito  = abs(amounts[-3][1])
                elif len(amounts) == 2:
                    val    = abs(amounts[-2][1])
                    x_val  = amounts[-2][0]
                    if (x_saldo - x_val) > w * 0.20:
                        debito, credito = val, 0.0   # columna DÉBITO (más alejada del SALDO)
                    else:
                        debito, credito = 0.0, val   # columna CRÉDITO (junto al SALDO)
                else:
                    debito, credito = 0.0, 0.0

                if debito == 0 and credito == 0: continue

                fecha_x   = next((x for x, t in line if t == fecha_str), 0)
                primer_monto_x = amounts[0][0]
                desc = " ".join(t for x, t in sorted(line)
                    if x > fecha_x and x < primer_monto_x - w * 0.01
                    and not re.match(r"^[\d,\.]+$", t)
                    and t not in ("O", "0", "SUC", "NIT")).strip()

                rows.append({"comprobante": comprobante, "fecha": fecha, "descripcion": desc,
                             "debito": debito, "credito": credito, "saldo": saldo})
            return pd.DataFrame(rows) if rows else pd.DataFrame(
                columns=["comprobante","fecha","descripcion","debito","credito","saldo"])

        def parse_auxiliar():
            """Procesa auxiliar con PyMuPDF (sin binarios del sistema)."""
            all_rows = []
            aux_sample = []
            try:
                doc = fitz.open(pdf_aux_path)
                for pg, page in enumerate(doc, 1):
                    words = page.get_text("words")
                    if words:
                        w = page.rect.width
                        lines, cur, cur_y = [], [], None
                        for x0,y0,x1,y1,word,*_ in sorted(words, key=lambda r:(round(r[1],1),r[0])):
                            if cur_y is None or abs(y0-cur_y)<=4:
                                cur.append((x0,word)); cur_y = cur_y or y0
                            else:
                                if cur: lines.append(sorted(cur))
                                cur, cur_y = [(x0,word)], y0
                        if cur: lines.append(sorted(cur))
                    else:
                        lines, w = _ocr_page(page, dpi=200)
                    if pg == 1:
                        aux_sample = [" | ".join(t for _, t in ln) for ln in lines[:8]]
                    df_p = _parse_informe_fitz(lines, w)
                    if not df_p.empty:
                        all_rows.append(df_p)
                doc.close()
            except Exception as e:
                aux_sample = [f"ERROR: {e}"]

            st.session_state["aux_sample"] = aux_sample
            if not all_rows:
                return pd.DataFrame(
                    columns=["comprobante","fecha","descripcion","debito","credito","saldo"]), None
            return pd.concat(all_rows, ignore_index=True), None

        # ── Matching ──────────────────────────────────────────────────────────
        def reconciliar(df_ext, df_inf, diferencia_saldos):
            # Índice de extracto: (fecha, monto_redondeado_100, tipo) → lista de índices
            # Redondear a $100 para absorber diferencias de centavos del OCR
            ext_idx = {}
            for i, row in df_ext.iterrows():
                for monto, tipo in [(row.debito,"D"),(row.credito,"C")]:
                    if monto > 0:
                        key = (row.fecha, round(monto / 100) * 100, tipo)
                        ext_idx.setdefault(key, []).append(i)

            matched_ext, matched_inf = set(), set()
            for i, row in df_inf.iterrows():
                for monto, tipo_ext in [(row.debito,"C"),(row.credito,"D")]:
                    if monto <= 0: continue
                    found = False
                    # ±10 días de tolerancia fecha, ±1% de tolerancia monto
                    tol_amt = max(round(monto * 0.01 / 100) * 100, 1000)
                    for day_d in [0,-1,-2,-3,-4,-5,-6,-7,-8,-9,-10,1,2,3,4,5]:
                        fecha_try = row.fecha + timedelta(days=day_d)
                        for amt_step in range(0, int(tol_amt) + 100, 100):
                            for sign in ([0] if amt_step == 0 else [amt_step, -amt_step]):
                                key = (fecha_try, round((monto + sign) / 100) * 100, tipo_ext)
                                if key in ext_idx and ext_idx[key]:
                                    matched_ext.add(ext_idx[key].pop(0))
                                    matched_inf.add(i)
                                    found = True; break
                            if found: break
                        if found: break

            _e = lambda lst, cols: pd.DataFrame(lst) if lst else pd.DataFrame(columns=cols)
            emp = lambda cols: pd.DataFrame(columns=cols)

            if round(abs(diferencia_saldos)) == 0:
                return (emp(["fecha","beneficiario","comprobante","monto"]),
                        emp(["fecha","descripcion","monto"]),
                        emp(["fecha","descripcion","monto"]),
                        emp(["fecha","descripcion","comprobante","monto"]),
                        True)

            # Devolver TODAS las partidas sin cruzar (no solo la más cercana al target)
            notas_cargo, notas_abono = [], []
            for i, row in df_ext.iterrows():
                if i in matched_ext: continue
                if row.debito > 0:
                    notas_cargo.append({"fecha": row.fecha, "descripcion": row.descripcion,
                                        "monto": row.debito})
                if row.credito > 0:
                    notas_abono.append({"fecha": row.fecha, "descripcion": row.descripcion,
                                        "monto": row.credito})

            cheques_pend, ingresos_pend = [], []
            for i, row in df_inf.iterrows():
                if i in matched_inf: continue
                if row.credito > 0:
                    cheques_pend.append({"fecha": row.fecha, "beneficiario": row.descripcion,
                                         "comprobante": row.comprobante, "monto": row.credito})
                if row.debito > 0:
                    ingresos_pend.append({"fecha": row.fecha, "descripcion": row.descripcion,
                                          "comprobante": row.comprobante, "monto": row.debito})

            # Ordenar por monto desc para que lo más relevante aparezca primero
            notas_cargo.sort(  key=lambda x: x["monto"], reverse=True)
            notas_abono.sort(  key=lambda x: x["monto"], reverse=True)
            cheques_pend.sort( key=lambda x: x["monto"], reverse=True)
            ingresos_pend.sort(key=lambda x: x["monto"], reverse=True)

            return (
                _e(cheques_pend,  ["fecha","beneficiario","comprobante","monto"]),
                _e(notas_cargo,   ["fecha","descripcion","monto"]),
                _e(notas_abono,   ["fecha","descripcion","monto"]),
                _e(ingresos_pend, ["fecha","descripcion","comprobante","monto"]),
                True,
            )

        def _buscar_partida(df_ext, df_inf, diferencia_saldos):
            """
            Detecta la partida sobrante usando un pool/contador:
            cada transacción del extracto puede absorber exactamente UNA entrada
            del auxiliar. Si el auxiliar tiene dos veces el mismo monto pero el
            banco solo una, la segunda queda sin par → es la partida extra.
            """
            from collections import defaultdict

            if df_ext.empty or df_inf.empty:
                return None

            target  = abs(diferencia_saldos)
            tol_dif = max(target * 0.15, 5_000)

            # Redondear a miles para absorber ruido OCR (±$500 queda en el mismo bucket)
            def _key(v):
                return round(v / 1_000) * 1_000

            def _build_pool(df):
                pool = defaultdict(int)
                for _, r in df.iterrows():
                    if r.debito  > 0: pool[_key(r.debito)]  += 1
                    if r.credito > 0: pool[_key(r.credito)] += 1
                return pool

            def _consume(pool, monto, tol_match=5_000):
                """Consume el bucket más cercano dentro de tolerancia. Devuelve True si encontró."""
                m_key = _key(monto)
                best_key, best_dist = None, float("inf")
                for k, cnt in pool.items():
                    if cnt > 0:
                        d = abs(k - m_key)
                        if d <= tol_match and d < best_dist:
                            best_dist, best_key = d, k
                if best_key is not None:
                    pool[best_key] -= 1
                    return True
                return False

            if diferencia_saldos < 0:
                # Banco < Contabilidad: el auxiliar tiene un CRÉDITO (abono) extra
                # que el banco no registró. Ese abono infla el saldo contable.
                # Comparamos CRÉDITOs del auxiliar contra CRÉDITOs del extracto.
                ext_c_pool = defaultdict(int)
                for _, r in df_ext.iterrows():
                    if r.credito > 0: ext_c_pool[_key(r.credito)] += 1

                candidatos = []
                for _, row in df_inf.iterrows():
                    monto = row.credito          # solo CRÉDITO en auxiliar
                    if monto < 1_000: continue
                    if _consume(ext_c_pool, monto):
                        pass  # tiene par en el banco → ok
                    else:
                        dist = abs(monto - target)
                        if dist <= tol_dif:
                            desc = getattr(row, "descripcion", "")
                            comp = getattr(row, "comprobante", "")
                            candidatos.append({
                                "origen":      "CRÉDITO en auxiliar sin respaldo en extracto bancario",
                                "fecha":       row.fecha,
                                "descripcion": f"{comp} {desc}".strip(),
                                "monto":       monto,
                                "tipo":        "CRÉDITO",
                                "dist_target": dist,
                            })
            else:
                # Banco > Contabilidad: el extracto tiene un CRÉDITO extra que
                # el auxiliar no registró, inflando el saldo bancario.
                aux_c_pool = defaultdict(int)
                for _, r in df_inf.iterrows():
                    if r.credito > 0: aux_c_pool[_key(r.credito)] += 1

                candidatos = []
                for _, row in df_ext.iterrows():
                    monto = row.credito          # solo CRÉDITO en extracto
                    if monto < 1_000: continue
                    if _consume(aux_c_pool, monto):
                        pass
                    else:
                        dist = abs(monto - target)
                        if dist <= tol_dif:
                            candidatos.append({
                                "origen":      "CRÉDITO en extracto bancario sin registrar en auxiliar",
                                "fecha":       row.fecha,
                                "descripcion": row.descripcion,
                                "monto":       monto,
                                "tipo":        "CRÉDITO",
                                "dist_target": dist,
                            })

            if not candidatos:
                return None
            candidatos.sort(key=lambda x: x["dist_target"])
            return candidatos[0]

        # ── Excel output ──────────────────────────────────────────────────────
        _HDR  = PatternFill("solid", fgColor="1F3864")
        _YELL = PatternFill("solid", fgColor="FFF2CC")
        _RED  = PatternFill("solid", fgColor="FCE4D6")
        _GRN  = PatternFill("solid", fgColor="E2EFDA")
        _GREY = PatternFill("solid", fgColor="F2F2F2")
        _LGRY = PatternFill("solid", fgColor="D9D9D9")
        _THIN = Side(style="thin", color="CCCCCC")
        _BRD  = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
        NCOLS = 5

        def _merge_row(ws, r, text, fill, font_kw=None, height=None):
            ws.merge_cells(f"A{r}:{get_column_letter(NCOLS)}{r}")
            c = ws[f"A{r}"]
            c.value = text
            c.fill = fill
            c.alignment = Alignment(horizontal="left", vertical="center")
            kw = {"bold": True, "size": 11, "color": "1F3864"}
            if font_kw: kw.update(font_kw)
            c.font = Font(**kw)
            if height: ws.row_dimensions[r].height = height

        def _kv(ws, r, label, valor, fill_v=None):
            """Fila de clave-valor: col A = etiqueta, col B = valor."""
            ca = ws.cell(row=r, column=1, value=label)
            ca.font = Font(bold=True, size=10)
            ca.border = _BRD
            cv = ws.cell(row=r, column=2, value=valor)
            cv.border = _BRD
            cv.alignment = Alignment(horizontal="right", vertical="center")
            cv.font = Font(size=10)
            if fill_v: cv.fill = fill_v

        def _hrow(ws, r, cols):
            for c, v in enumerate(cols, 1):
                cell = ws.cell(row=r, column=c, value=v)
                cell.font = Font(bold=True, color="FFFFFF", size=9)
                cell.fill = _HDR
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = _BRD

        def _drow(ws, r, vals, fill=None):
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=r, column=c, value=v)
                if fill: cell.fill = fill
                cell.border = _BRD
                cell.alignment = Alignment(vertical="center")

        def _total_row(ws, r, label, total):
            ws.merge_cells(f"A{r}:D{r}")
            c = ws[f"A{r}"]
            c.value = label
            c.font = Font(bold=True, size=10)
            c.fill = _LGRY
            c.border = _BRD
            c.alignment = Alignment(horizontal="right")
            cv = ws.cell(row=r, column=5, value=total)
            cv.font = Font(bold=True, size=10)
            cv.fill = _LGRY
            cv.border = _BRD
            cv.alignment = Alignment(horizontal="right")

        def build_excel(df_ext, df_inf, saldo_ext, saldo_cont, nombre_pdf, partida_faltante):
            wb = Workbook()
            ws = wb.active; ws.title = "Conciliacion"
            for col, cw in zip("AB", [38, 26]):
                ws.column_dimensions[col].width = cw

            diferencia = round((saldo_ext or 0) - (saldo_cont or 0), 2)

            # Título
            ws.merge_cells("A1:B1")
            c = ws["A1"]
            c.value = "CONCILIACIÓN BANCARIA"
            c.font = Font(bold=True, size=14, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="1F3864")
            c.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 30

            ws.merge_cells("A2:B2")
            c2 = ws["A2"]
            c2.value = f"{nombre_pdf}   |   {datetime.now().strftime('%Y-%m-%d %H:%M')}"
            c2.font = Font(size=9, color="666666")
            c2.fill = PatternFill("solid", fgColor="F2F2F2")
            c2.alignment = Alignment(horizontal="center")

            _kv(ws, 4, "SALDO EXTRACTO BANCARIO", _fmt_cop(saldo_ext),  _GRN)
            _kv(ws, 5, "SALDO CONTABILIDAD",       _fmt_cop(saldo_cont), _GRN)
            _kv(ws, 6, "DIFERENCIA",               _fmt_cop(diferencia),
                _GRN if diferencia == 0 else _RED)

            if partida_faltante:
                ws.merge_cells("A8:B8")
                lbl = ws["A8"]
                lbl.value = "PARTIDA FALTANTE IDENTIFICADA"
                lbl.font = Font(bold=True, size=11, color="7F4F00")
                lbl.fill = _YELL
                lbl.alignment = Alignment(horizontal="center", vertical="center")
                ws.row_dimensions[8].height = 22

                _kv(ws, 9,  "Fecha",       str(partida_faltante["fecha"]),       _YELL)
                _kv(ws, 10, "Descripción", partida_faltante["descripcion"],      _YELL)
                _kv(ws, 11, "Monto",       _fmt_cop(partida_faltante["monto"]),  _YELL)
                _kv(ws, 12, "Tipo",        partida_faltante["tipo"],             _YELL)
                _kv(ws, 13, "Origen",      partida_faltante["origen"],           _YELL)

            # Hoja Extracto — última fila usa saldo_ext oficial
            ws2 = wb.create_sheet("Extracto")
            for col, cw in zip("ABCDEF",[12,8,45,22,22,26]): ws2.column_dimensions[col].width = cw
            _hrow(ws2, 1, ["Fecha","Día","Descripción","Débito","Crédito","Saldo"])
            ext_rows = list(df_ext.iterrows())
            for i, (_, row) in enumerate(ext_rows, 2):
                fill = _GRN if i%2==0 else _GREY
                es_ultima = (i == len(ext_rows) + 1)
                saldo_fila = saldo_ext if (es_ultima and saldo_ext) else row.saldo
                _drow(ws2, i, [str(row.fecha), row.dia, row.descripcion,
                    _fmt_cop(row.debito) if row.debito>0 else "",
                    _fmt_cop(row.credito) if row.credito>0 else "",
                    _fmt_cop(saldo_fila)], fill)

            # Hoja Auxiliar
            ws3 = wb.create_sheet("Auxiliar")
            for col, cw in zip("ABCDEF",[18,12,40,22,22,26]): ws3.column_dimensions[col].width = cw
            _hrow(ws3, 1, ["Comprobante","Fecha","Descripción","Débito","Crédito","Saldo"])
            for i,(_, row) in enumerate(df_inf.iterrows(), 2):
                fill = _GRN if i%2==0 else _GREY
                _drow(ws3, i, [row.comprobante, str(row.fecha), row.descripcion,
                    _fmt_cop(row.debito) if row.debito>0 else "",
                    _fmt_cop(row.credito) if row.credito>0 else "",
                    _fmt_cop(row.saldo)], fill)

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            return buf, diferencia

        # ── Ejecutar ──────────────────────────────────────────────────────────
        meses_str = {"ene":1,"feb":2,"mar":3,"abr":4,"may":5,"jun":6,
                     "jul":7,"ago":8,"sep":9,"oct":10,"nov":11,"dic":12}
        # Intenta "01-MAR-2026", luego "MAR.2026" o "MAR 2026"
        m = re.search(r"(\d{1,2})[\.\-]([A-Za-z]{3,4})[\.\-](\d{4})", uploaded_ext.name)
        if m:
            ano = int(m.group(3))
            mes = meses_str.get(m.group(2).lower()[:3], 1)
        else:
            m2 = re.search(r"\b([A-Za-z]{3,4})[\.\-\s](\d{4})\b", uploaded_ext.name)
            if m2:
                ano = int(m2.group(2))
                mes = meses_str.get(m2.group(1).lower()[:3], date.today().month)
            else:
                ano, mes = date.today().year, date.today().month

        # Procesar extracto
        df_ext, _images_ext = parse_extracto(ano, mes)

        # 1) Buscar "Saldo Final" en OCR de últimas páginas (más confiable)
        saldo_ext = _find_saldo_fitz(pdf_ext_path)

        # 2) Fallback: saldo máximo de la tabla OCR (el saldo corriente crece hasta el final)
        if saldo_ext is None and not df_ext.empty:
            _ext_saldos = df_ext["saldo"].dropna()
            if not _ext_saldos.empty:
                saldo_ext = float(_ext_saldos.max())

        # Procesar auxiliar — pdfplumber lee texto directo, sin OCR
        df_inf, _images_aux = parse_auxiliar()
        _inf_saldos = df_inf["saldo"].dropna() if not df_inf.empty else pd.Series(dtype=float)
        saldo_cont = float(_inf_saldos.iloc[-1]) if not _inf_saldos.empty else None

        diferencia_saldos = round((saldo_ext or 0) - (saldo_cont or 0), 2)
        partida_faltante  = _buscar_partida(df_ext, df_inf, diferencia_saldos)

        excel_buf, diferencia = build_excel(
            df_ext, df_inf, saldo_ext, saldo_cont,
            uploaded_ext.name, partida_faltante,
        )

        os.unlink(pdf_ext_path)
        os.unlink(pdf_aux_path)

    # ── Resultado ─────────────────────────────────────────────────────────────
    col1, col2, col3 = st.columns(3)
    col1.metric("Saldo Extracto",      f"${saldo_ext:,.0f}"  if saldo_ext  else "—")
    col2.metric("Saldo Contabilidad",  f"${saldo_cont:,.0f}" if saldo_cont else "—")
    col3.metric("Diferencia",          f"${diferencia:,.0f}" if diferencia is not None else "—",
                delta_color="off" if diferencia==0 else "inverse")

    n_ext = len(df_ext) if not df_ext.empty else 0
    n_aux = len(df_inf) if not df_inf.empty else 0
    st.caption(f"Extracto: **{n_ext} movimientos** extraídos   |   Auxiliar: **{n_aux} movimientos** extraídos")

    if diferencia == 0:
        st.success("✅ Conciliación completa — los saldos cuadran perfectamente.")
    else:
        st.error(f"⚠️ Diferencia de **${diferencia:,.0f}**")

        if partida_faltante:
            origen = partida_faltante["origen"]
            st.warning(
                f"**Posible partida faltante — {origen}:**\n\n"
                f"- Fecha: {partida_faltante['fecha']}\n"
                f"- Descripción: {partida_faltante['descripcion']}\n"
                f"- Monto: **${partida_faltante['monto']:,.0f}** ({partida_faltante['tipo']})"
            )

    nombre_salida = uploaded_ext.name.replace(".pdf", "_conciliacion.xlsx")
    st.download_button(
        label="⬇️ Descargar Excel de conciliación",
        data=excel_buf,
        file_name=nombre_salida,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
    )

    with st.expander("Ver movimientos del extracto"):
        cols_ext = [c for c in ["fecha","dia","descripcion","debito","credito","saldo"] if c in df_ext.columns]
        if cols_ext:
            st.dataframe(df_ext[cols_ext], use_container_width=True)
        else:
            st.warning("pdfplumber no extrajo filas del extracto — ver texto de debug abajo.")
        debug = st.session_state.get("ext_debug", [])
        if debug:
            modo = debug[0][2] if debug else "?"
            st.caption(f"Modo: {'pdfplumber ✓' if modo=='pdf' else 'OCR'} — "
                       "Filas/página: " + " | ".join(f"P{p}:{n}" for p, n, _ in debug))
        sample = st.session_state.get("plumber_sample", [])
        if sample:
            st.caption("**Texto que leyó pdfplumber del extracto (pág 1) — cópiame esto:**")
            for ln in sample:
                st.code(ln)

    with st.expander("Ver asientos del contador"):
        cols_inf = [c for c in ["comprobante","fecha","descripcion","debito","credito","saldo"] if c in df_inf.columns]
        if cols_inf:
            st.dataframe(df_inf[cols_inf], use_container_width=True)
        else:
            st.warning("pdfplumber no extrajo filas del auxiliar.")
        aux_sample = st.session_state.get("aux_sample", [])
        if aux_sample:
            st.caption("**Texto que leyó pdfplumber del auxiliar (pág 1) — cópiame esto:**")
            for ln in aux_sample:
                st.code(ln)
